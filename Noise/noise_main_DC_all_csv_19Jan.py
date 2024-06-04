# -*- coding: utf-8 -*-


"""Initial creation date Tue Nov 13 14:30:27 2018
This code will work for DC method using 2450 Keithely 


   @author: Dasharath Adhikari
   @Editior: Nitin Kumar
   Important:
       you need to specify
       1) location of experimental data files = 'path_raw'
   
"""


import os
from scipy import signal
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib as mpl
import openpyxl
import xlrd
from openpyxl import load_workbook

plt.rcParams['agg.path.chunksize']=10000  # Use the default value
mpl.rcParams['xtick.direction'] = 'in'
mpl.rcParams['ytick.direction'] = 'in'
mpl.rcParams['xtick.top'] = True
mpl.rcParams['ytick.right'] = True
plt.rcParams['axes.linewidth'] = 2

# Update the font settings
font = {'family' : 'Times New Roman',
        'weight' : 'normal',
        'size'   : 18}  
plt.rc('font', **font)
plt.rc('lines', linewidth=2)
plt.rcParams['axes.labelweight'] = 'bold'

from low_pass_kaiser_window_design import  three_stage_decimation
from psd_slope_alpha import alpha_calculate


'''Change file path'''


base_directory = r"F:\Personal Data\NbO2\Size dependent study\Line 5\New folder\16 Min - Copy"

path_csv = base_directory
path_text = os.path.join(base_directory, "text_file")
path_detrend = os.path.join(base_directory, "detrend_file")
path_DTS = os.path.join(base_directory, "dts_file")
path_DTS_padding = os.path.join(base_directory, "dts_padding_file")
path_psd = os.path.join(base_directory, "psd_file")
path_npsd = os.path.join(base_directory, "psdn_file")
plots_folder = os.path.join(base_directory, "plots")

plots_folder = os.path.join(path_csv, "plots")
if not os.path.exists(plots_folder):
    os.makedirs(plots_folder)
    
# Create directories if they don't exist
os.makedirs(path_text, exist_ok=True)
os.makedirs(path_detrend, exist_ok=True)
os.makedirs(path_DTS, exist_ok=True)
os.makedirs(path_DTS_padding, exist_ok=True)
os.makedirs(path_psd, exist_ok=True)
os.makedirs(path_npsd, exist_ok=True)
os.makedirs(plots_folder, exist_ok=True)
    
    
for file in os.listdir(path_csv):
    print(file)


'''Set range of file numbers to import files'''
# Specify the range of file numbers you want to analyze

start_file_number = 78  # Change this to the starting file number
end_file_number = 132   # Change this to the ending file number

for filenumber in range(start_file_number, end_file_number + 1):
    temperature = str(filenumber).zfill(3)
    

    # Construct the file format for the current file
    file_format = 'it_' + temperature + '.csv'
    
    
    '''Get time series from the csv file'''

    # Read CSV file as a Pandas DataFrame
    dataframe = pd.read_csv(path_csv + os.sep + file_format, delimiter=',',
                                header=None,  skiprows=9, skipfooter=0, nrows=None, low_memory=False, encoding_errors="ignore")
    
    i = dataframe.iloc[:, 1].to_numpy()
    voltage_applied = dataframe.iloc[9, 14]
   #i = dataframe.iloc[:, 14].to_numpy()
    bias_input = voltage_applied
    iavg=i.mean()
    print('Average current: ', i.mean())
    
    
    '''Make time array based on the sampling frequency'''
    fs = 500.0 # sampling frequency
    t_list = []
    t0 = 0.0
    for n in range(len(i)):    
        t_list.append(t0)
        dt = 1/fs
        t0 += dt 
    t_list = np.asarray(t_list) 
    
    print('Creating a text file of current time series...')
    text_file_format = 'it_'+ temperature+'.txt'
    file = open(path_text + os.sep + text_file_format, 'w')
    file.write('current(A)'+str(filenumber).zfill(3) + '\t' + 'time(s)'+str(filenumber).zfill(3) + '\n')
    for m in range(len(t_list)):
        file.write(str(i[m]) +'\t'+ str(t_list[m]) +'\n')  
    file.close()
    print('Creating text file completed')
    
    print('Calculating PSD from the current time series....')
    
    # t = time
    # i = current 
    i, t = np.loadtxt(path_text + os.sep + text_file_format, skiprows=1,
                                usecols=[0, 1], unpack=True)
    
    idetrend = signal.detrend(i, axis=-1, type='linear', bp=[0,len(i)])
    #idetrend /= iavg
    #idetrend = signal.detrend(i, axis=-1, type='constant', bp=[0,len(i)]) # Can choose linear/constant
    
    def detrend_file(_write=False):
        if _write:
            # creates the text file of the detrended time series
            detrend_file_name = 'detrend_'+text_file_format
            FileToWrite = open(path_detrend + os.sep + detrend_file_name, 'w')
            FileToWrite.write("t"+str(filenumber).zfill(3) + "\t" + "i_detrend"+str(filenumber).zfill(3) + "\n" )
            for idx, times in enumerate(t):
                FileToWrite.write(str(round(times, 5)) +"\t"+ str(idetrend[idx]) + "\n")       
            FileToWrite.close()
    detrend_file(_write=True)
    
    # plot the time series if plot is set True
    '''PLot time series before and after detrending'''
    def plot_signal_time_series(time, i, idetrend, plot=False):
        if plot:    
            # Plotting time series
            fig, ax = plt.subplots(1, 2, figsize=(12, 6), dpi=100)
            ax[0].plot(time, i, c="r", label="Before detrending")
            ax[0].set_xlabel(r't (s)',  fontsize=18)
            ax[0].set_ylabel(r'$\delta$I (A)', fontsize=18)
            ax[0].legend(loc='best')
            ax[1].plot(time, idetrend, c="r", label="After detrending")
            ax[1].set_xlabel(r't (s)',  fontsize=18)
            ax[1].legend(loc='best')
            
            file_name_1 = 'time_series_plot_' + temperature + '.png'
            plt.savefig(os.path.join(plots_folder, file_name_1), dpi=300, bbox_inches='tight')
            plt.show()
    plot_signal_time_series(t, i, idetrend, plot=True) # Can set True/False
    
    
    
    ''' Decimation'''
    
    fs = 500.0  # data sampling frequency
    D1 = 4.0  # first stage decimation factor
    D2 = 2.0  # second stage decimation factor
    D3 = 2.0  # third stage decimation factor
    decimation_factor = D1*D2*D3
    f_pass_final = (0.5/decimation_factor) * fs
    print(r'f_sampling = {} Hz, D1= {}, D2= {}, and D3= {}'.format(fs, D1, D2, D3))
    #i = idetrend
    i = three_stage_decimation(idetrend, fs, f_pass_final, D1, D2, D3)
    t_new = np.linspace(start=0, stop=int(t.max()),
                        num=len(i), endpoint=True)
    
    def plot_decimated_signal_time_series(time, i, plot=False):
        if plot:
            # Plotting time series
            fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=150)         
            ax.plot(time, i, c="r", label="Signal")
            ax.set_xlabel(r't (s)',  fontsize=18)
            ax.set_ylabel(r'$\delta$I (A)', fontsize=18)
            ax.set_title('Signal after decimation')
            ax.legend(loc='best')
            
            file_name_2 = 'signal_after_decimation_' + temperature + '.png'
            plt.savefig(os.path.join(plots_folder, file_name_2), dpi=300, bbox_inches='tight')
            plt.show()
    plot_decimated_signal_time_series(t_new, i, plot=True)
    
    def dts_file(_write=False):
        if _write:
            # creates the text file of calculated spectral density
            #dts_file_name = 'DTS_'+file_to_read
            dts_file_name = 'DTS_'+text_file_format
            FileToWrite = open(path_DTS + os.sep + dts_file_name, 'w')
            FileToWrite.write("t_new"+str(filenumber).zfill(3) + "\t" + "i" +str(filenumber).zfill(3) +"\n" )
            for idx, times in enumerate(t_new):
                FileToWrite.write(str(round(times, 5)) +"\t"+ str(i[idx]) + "\n")       
            FileToWrite.close()
    dts_file(_write=True)
    
    n = len(i)
    m = len(t_new)
    
    # zero padding vx and time in order to make data, power of 2
    
    n1 = int(np.log2(n))
    n1 = np.power(2, n1 + 1)
    m1 = int(np.log2(m))
    m1 = np.power(2, m1 + 1)
    padding = np.zeros((n1 - n))
    ipad = np.concatenate((i, padding))
    tpad = np.concatenate((t_new, padding))
    #vy = np.concatenate((vy, padding))
    
    def dts_padding_file(_write=False):
        if _write:
            # creates the text file of padded decimates time series
            dts_padding_file_name = 'dts_padding_'+text_file_format
            FileToWrite = open(path_DTS_padding + os.sep + dts_padding_file_name, 'w')
            FileToWrite.write("t_new" + "\t" + "i" "\n" )
            for idx, times in enumerate(tpad):
                FileToWrite.write(str(round(times, 5)) +"\t"+ str(ipad[idx]) + "\n")       
            FileToWrite.close()
    dts_padding_file(_write=True)
    
    nperseg = 1024*4
    
    def power_spectrum_density(x, spectrum=False):
        """
        Caculate spectral density using Welch periodogram method.
        x : array_like, time series of measurement values
        fs :sampling frequency of the x time series in units of Hz
        window : desired window, defaults to ‘hanning’
        nperseg : length of each segment. Defaults to 256
        noverlap: number of points to overlap between segments
        If None, noverlap = nperseg / 2
        nfft : Length of the FFT used, if a zero padded FFT is desired
        If None, the FFT length is nperseg. Defaults to None.
        detrend: specifies how to detrend each segment. defaults to ‘constant’.
        return_onesided : bool,if True, return a one-sided spectrum for real data
        if False return a two-sided spectrum
        scaling : { ‘density’, ‘spectrum’ }, optional
        """
        if spectrum:
            f, S = signal.welch(x, fs=fs/(D1*D2*D3), window='hann',
                                nperseg=nperseg, noverlap=0.5*nperseg, nfft=None,
                                detrend='constant', return_onesided=True,
                                scaling='spectrum', axis=-1)
        else:
            f, S = signal.welch(x, fs=fs/(D1*D2*D3), window='hann',
                                nperseg=nperseg, noverlap=0.5*nperseg, nfft=None,
                                detrend='constant', return_onesided=True,
                                scaling='density', axis=-1)
        return f, S
    
    # spectral density estimation from vx data
    fx, Sx = power_spectrum_density(ipad)
    
    # not selecting zero frequency component
    fx = fx[1:]
    Sx = Sx[1:]
    
    def psd_plot(plot=True):
        if plot:
            fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=150)
            ax.loglog(fx, Sx, c="r", lw=2.0, label="PSD (unnormalized)")
            ax.set_xlabel(r'f [Hz]', fontsize=18)
            ax.set_ylabel(r'S$_{I}$ [$\frac{I^{2}}{Hz}$]', fontsize=18)

            ax.set_title('Power Spectral Density')
            ax.legend(loc='best')
            
            file_name_3 = 'psd_plot_' + temperature + '.png'
            plt.savefig(os.path.join(plots_folder, file_name_3), dpi=300, bbox_inches='tight')
            plt.show()
    psd_plot(plot=True)
    
    def psd_file(_write=True):
        if _write:
            # creates the text file of calculated spectral density
            psd_file_format = 'psd_'+temperature+'.txt'
            FileToWrite = open(path_psd + os.sep + psd_file_format, 'w')
            FileToWrite.write("f"+str(filenumber).zfill(3) + "\t" + "Sx"+str(filenumber).zfill(3) + "\n" )
            for idx, freq in enumerate(fx):
                FileToWrite.write(str(round(freq, 5)) +"\t"+ str(Sx[idx]) + "\n")       
            FileToWrite.close()
    
    psd_file(_write=True) 
    # Normalize PSD by dividing by the square of the average current
    Sx_normalized = Sx / (iavg**2)
    # Remove the zero frequency component
    fx = fx[1:]
    Sx_normalized = Sx_normalized[1:]
    
    def psd_plot(plot=True):
        if plot:
            fig, ax = plt.subplots(1, 1, figsize=(10, 8), dpi=150)
            ax.loglog(fx, Sx_normalized, c="r", lw=2.0, label="Normalized PSD")
            ax.set_xlabel(r'f [Hz]', fontsize=18)
            ax.set_ylabel(r'S$_{I}$/I$_{avg}^{2}$', fontsize=18)
            ax.set_title('Normalized Power Spectral Density')
            ax.legend(loc='best')
            
            file_name_5 = 'normalized_psd_plot_' + temperature + '.png'
            plt.savefig(os.path.join(plots_folder, file_name_5), dpi=300, bbox_inches='tight')
            plt.show()
    
    psd_plot(plot=True)
    def psd_file(_write=True):
        if _write:
            # creates the text file of calculated spectral density
            psdn_file_format = 'psdn_'+temperature+'.txt'
            FileToWrite = open(path_npsd + os.sep + psdn_file_format, 'w')
            FileToWrite.write("fn"+str(filenumber).zfill(3) + "\t" + "Sx_normalized"+str(filenumber).zfill(3) + "\n" )
            for idx, freq in enumerate(fx):
                FileToWrite.write(str(round(freq, 5)) +"\t"+ str(Sx_normalized[idx]) + "\n")       
            FileToWrite.close()
    
    psd_file(_write=True) 


    # alpha_calculate gives frequency exponent, slope 
    
    f_range = [0.01, 2]
    
    # f_range = [0.01, 8.0]
    
    slope, intercept, std_err = alpha_calculate(f=fx, S=Sx, f_range=f_range)
    
    f_new = np.linspace(0.001, 1.0, 256)
    f_new = np.log10(f_new)
    S_new = slope*f_new + intercept
    S_new = 10**S_new
    f_new = 10**f_new
    
    # Create the "plots" folder in the same directory as the CSV files
    csv_folder = path_csv  # Change this to your CSV files directory
    plots_folder = os.path.join(csv_folder, "plots")
    if not os.path.exists(plots_folder):
        os.makedirs(plots_folder)
    
    plt.close('all')
            
    #--------------------------------------------------------------------------
    #--------------------------------------------------------------------------
    ''' To export data into excel file'''
    from os.path import join
    from psd_slope_alpha import rsq
    
    path_NoiseData = path_csv
    
    excel_file_path = join(path_NoiseData, 'Noise_IV_Analysis.xlsx')

    # Try to open the workbook, create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(excel_file_path)
        
        # Load the existing workbook
    wb = load_workbook(join(path_NoiseData, 'Noise_IV_Analysis.xlsx'))
        
        # Check if the sheet 'NoiseData' already exists
    sheet_name = 'NoiseData'
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        # Create a new sheet
        sheet = wb.create_sheet(sheet_name, 0)
    
    
    # Your calculations
    foooi = (0.0039 ** slope) * (10 ** intercept)
    fooi = (0.01 ** slope) * (10 ** intercept)
    foi = (0.1 ** slope) * (10 ** intercept)
    fi = (1 ** slope) * (10 ** intercept)
    ''' Calculation for Normalized magnitudes'''
    
    nfoooi = (0.0039 ** slope) * (10 ** intercept)/(iavg**2)
    nfooi = (0.01 ** slope) * (10 ** intercept)/(iavg**2)
    nfoi = (0.1 ** slope) * (10 ** intercept)/(iavg**2)
    nfi = (1 ** slope) * (10 ** intercept)/(iavg**2)
    
    # Write header row
    header = ['File number', 'Avg of Fluc', 'Slope', 'Err in slope', 'R Square',
              'Noise at 3.9 mHz', 'Noise at 10 mHz', 'Noise at 100 mHz', 'Noise at 1 Hz',
              'Volt/Curr Supplied (V/A)','Normalized Noise at 3.9 mHz', 'Normalized Noise at 10 mHz', 'Normalized Noise at 100 mHz', 'Normalized Noise at 1 Hz']
    
    for col_num, value in enumerate(header, 1):
        sheet.cell(row=1, column=col_num, value=value)
    
    # Write data row
    data = [int(filenumber), iavg, slope, std_err, rsq, foooi, fooi, foi, fi, bias_input, nfoooi, nfooi, nfoi, nfi]
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=int(filenumber) + 1, column=col_num, value=value)
    
    # Save the workbook
    
    wb.save(join(path_NoiseData, 'Noise_IV_Analysis.xlsx'))
    
    print('Data Saved')

    

